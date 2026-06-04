import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from collections import defaultdict
import sys

# ============================================================
# CONFIG
# ============================================================
THIN = Side(style='thin')
MEDIUM = Side(style='medium')
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_MEDIUM = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

FONT_HEADER = Font(name='맑은 고딕', bold=True, size=10)
FONT_TITLE = Font(name='맑은 고딕', bold=True, size=12)
FONT_DATA = Font(name='맑은 고딕', size=9)
FONT_PCT = Font(name='맑은 고딕', size=9, color='FF0000')

FILL_HEADER = PatternFill('solid', fgColor='4472C4')
FILL_SUB_HEADER = PatternFill('solid', fgColor='D6E4F0')
FILL_SECTION = PatternFill('solid', fgColor='E2EFDA')
FILL_LIGHT_YELLOW = PatternFill('solid', fgColor='FFF2CC')
FILL_LIGHT_GRAY = PatternFill('solid', fgColor='F2F2F2')
FILL_GAP_POS = PatternFill('solid', fgColor='DAEEF3')
FILL_GAP_NEG = PatternFill('solid', fgColor='F2DCDB')

ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

# ============================================================
# 1. Parse ALL original data
# ============================================================
wb_src = openpyxl.load_workbook('/sessions/eloquent-laughing-bohr/mnt/sono-competitor-crawler/소노벨단양_6월프로모션_원본.xlsx', data_only=True)

def parse_sheet(ws):
    records = []
    cur_dong = cur_type = cur_view = None
    for r in range(9, ws.max_row + 1):
        c4 = ws.cell(row=r, column=4).value
        c5 = ws.cell(row=r, column=5).value
        c6 = ws.cell(row=r, column=6).value
        c8 = ws.cell(row=r, column=8).value
        if c4: cur_dong = str(c4).strip()
        if c5: cur_type = str(c5).strip()
        if c6: cur_view = str(c6).strip()
        if c8 is None: continue
        season = str(c8).strip()
        if '2026-06-05' in season or '06-05' in season or '06-06' in season:
            season = '현충일'
        elif '2026-07-16' in season or '07-16' in season:
            season = '7/16'
        
        def g(col): return ws.cell(row=r, column=col).value
        rec = {
            'dong': cur_dong, 'type': cur_type, 'view': cur_view, 'season': season,
            '정상가_기명': g(9), '정상가_무기명': g(10), '정상가_FIT': g(11),
            '정상가_조식': g(12), '정상가_오션': g(13), '정상가_레전드': g(14), '정상가_웰컴티': g(15),
            '스마트_기명': g(17), '스마트_무기명': g(18), '스마트_FIT': g(19),
            '배분가_기명': g(23), '배분가_기명_할인율': g(24),
            '배분가_무기명': g(25), '배분가_무기명_할인율': g(26),
            '배분가_FIT': g(27), '배분가_FIT_할인율': g(28),
            '배분가_조식': g(29), '배분가_조식_할인율': g(30),
            '배분가_오션': g(31), '배분가_오션_할인율': g(32),
            '배분가_레전드1': g(33), '배분가_레전드1_할인율': g(34),
            '배분가_레전드2': g(35), '배분가_레전드2_할인율': g(36),
            '배분가_웰컴1': g(37), '배분가_웰컴1_할인율': g(38),
            '배분가_웰컴2': g(39), '배분가_웰컴2_할인율': g(40),
            '상품1_기명': g(42), '상품1_무기명': g(43), '상품1_FIT': g(44),
            '상품1_수수료011': g(45), '상품1_입금가011': g(46),
            '상품1_수수료013': g(47), '상품1_입금가013': g(48),
            '상품2_기명': g(51), '상품2_무기명': g(52), '상품2_FIT': g(53),
            '상품2_수수료011': g(54), '상품2_입금가011': g(55),
            '상품2_수수료013': g(56), '상품2_입금가013': g(57),
        }
        records.append(rec)
    return records

data_6 = parse_sheet(wb_src['상품안_6월'])
data_7 = parse_sheet(wb_src['상품안_7~7.16월'])
data_ss = parse_sheet(wb_src['상품안_성수기'])

# ============================================================
# 2. Build MIN/MAX ranges per (dong, type, view)
# ============================================================
# For each room type, across all seasons in a period, compute MIN and MAX

def compute_min_max(records, fields):
    """Group by (dong, type, view) and compute min/max for given fields across all seasons."""
    groups = defaultdict(list)
    for r in records:
        key = (r['dong'], r['type'], r['view'])
        groups[key].append(r)
    
    result = {}
    for key, recs in groups.items():
        minmax = {}
        for f in fields:
            vals = [r[f] for r in recs if r[f] is not None and isinstance(r[f], (int, float))]
            if vals:
                minmax[f'{f}_min'] = min(vals)
                minmax[f'{f}_max'] = max(vals)
            else:
                minmax[f'{f}_min'] = None
                minmax[f'{f}_max'] = None
        # Also keep individual season data for reference
        minmax['seasons'] = {r['season']: r for r in recs}
        result[key] = minmax
    return result

PRICE_FIELDS = [
    '정상가_기명', '정상가_무기명', '정상가_FIT',
    '스마트_기명', '스마트_무기명', '스마트_FIT',
    '배분가_기명', '배분가_무기명', '배분가_FIT',
    '배분가_기명_할인율', '배분가_무기명_할인율', '배분가_FIT_할인율',
    '배분가_조식', '배분가_오션', '배분가_레전드1', '배분가_레전드2', '배분가_웰컴1', '배분가_웰컴2',
    '배분가_조식_할인율', '배분가_오션_할인율', '배분가_레전드1_할인율', '배분가_레전드2_할인율', '배분가_웰컴1_할인율', '배분가_웰컴2_할인율',
    '상품1_기명', '상품1_무기명', '상품1_FIT',
    '상품1_수수료011', '상품1_입금가011', '상품1_수수료013', '상품1_입금가013',
    '상품2_기명', '상품2_무기명', '상품2_FIT',
    '상품2_수수료011', '상품2_입금가011', '상품2_수수료013', '상품2_입금가013',
]

mm_6 = compute_min_max(data_6, PRICE_FIELDS)
mm_7 = compute_min_max(data_7, PRICE_FIELDS)
mm_ss = compute_min_max(data_ss, PRICE_FIELDS)

# ============================================================
# 3. Create proposal Excel
# ============================================================

def fmt_range(mn, mx):
    """Format as 'MIN ~ MAX' or single value if same."""
    if mn is None: return '-'
    if mn == mx or mx is None: return int(mn) if isinstance(mn, float) and mn == int(mn) else mn
    return f"{int(mn):,} ~ {int(mx):,}"

def fmt_pct_range(mn, mx):
    """Format percentage range."""
    if mn is None: return '-'
    if isinstance(mn, (int, float)):
        mn_p = round(mn * 100, 1)
        mx_p = round(mx * 100, 1) if mx else mn_p
        if mn_p == mx_p: return f"{mn_p}%"
        return f"{mn_p}% ~ {mx_p}%"
    return '-'

def safe_int(v):
    if v is None: return 0
    if isinstance(v, float): return int(v)
    return v

def create_proposal(mode, filename):
    """
    mode: '무기명' or '기명'
    Creates a proposal Excel with MIN~MAX ranges.
    """
    wb = openpyxl.Workbook()
    
    # ---- Sheet 1: 요금 요약 ----
    ws = wb.active
    ws.title = '요금요약'
    
    # Title
    ws.merge_cells('A1:T1')
    ws['A1'] = f'■ 소노벨 단양 6월 판매 프로모션 요금(안) — {mode} 기준'
    ws['A1'].font = FONT_TITLE
    ws['A1'].alignment = ALIGN_LEFT
    
    ws.merge_cells('A2:T2')
    ws['A2'] = f'시즌별 가격이 다를 때 MIN~MAX 범위로 표시 | 회원({mode}) 기준'
    ws['A2'].font = Font(name='맑은 고딕', size=9, color='808080')
    
    # Define periods and their data
    periods = [
        ('6월 (6/1~6/30)', mm_6, data_6),
        ('7/1~7/16', mm_7, data_7),
        ('성수기 (7/17~8/22)', mm_ss, data_ss),
    ]
    
    # Room order
    room_order_west = [
        ('WEST', '패밀리', '스탠다드뷰'),
        ('WEST', '패밀리', '레이크뷰'),
        ('WEST', '스위트', '스탠다드뷰'),
        ('WEST', '스위트', '레이크뷰'),
    ]
    room_order_east = [
        ('EAST', '패밀리', '스탠다드뷰'),
        ('EAST', '패밀리', '탑스탠다드뷰'),
        ('EAST', '패밀리', '레이크뷰'),
        ('EAST', '패밀리', '탑레이크뷰'),
        ('EAST', '스위트', '스탠다드뷰'),
        ('EAST', '스위트', '탑스탠다드뷰'),
        ('EAST', '스위트', '레이크뷰'),
        ('EAST', '스위트', '탑레이크뷰'),
    ]
    all_rooms = room_order_west + room_order_east
    
    # Member key based on mode
    mem_key = '무기명' if mode == '무기명' else '기명'
    
    row = 4
    for period_name, mm_data, raw_data in periods:
        # Period header
        ws.merge_cells(f'A{row}:T{row}')
        ws[f'A{row}'] = period_name
        ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
        ws[f'A{row}'].fill = PatternFill('solid', fgColor='2F5496')
        ws[f'A{row}'].alignment = ALIGN_LEFT
        row += 1
        
        # --- Section: 객실 정상가/배분가/GAP ---
        ws.merge_cells(f'A{row}:T{row}')
        ws[f'A{row}'] = '▶ 객실 요금'
        ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, size=10)
        ws[f'A{row}'].fill = FILL_SECTION
        row += 1
        
        # Column headers for 객실
        headers = ['동', '타입', '뷰', 
                   f'정상가\n회원({mem_key})', '정상가\nFIT',
                   f'스마트\n회원({mem_key})', '스마트\nFIT',
                   f'배분가\n회원({mem_key})', f'배분가\n할인율', '배분가\nFIT', '배분가\nFIT할인율',
                   'GAP\n(FIT-회원)', 'GAP율']
        
        for ci, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='맑은 고딕', bold=True, size=8, color='FFFFFF')
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1
        
        # Data rows
        for dong, rtype, view in all_rooms:
            key = (dong, rtype, view)
            if key not in mm_data:
                continue
            mm = mm_data[key]
            
            # Get season data for per-season display
            seasons = mm.get('seasons', {})
            
            # Compute min/max for the mode
            正_회원_min = mm.get(f'정상가_{mem_key}_min')
            正_회원_max = mm.get(f'정상가_{mem_key}_max')
            正_FIT_min = mm.get('정상가_FIT_min')
            正_FIT_max = mm.get('정상가_FIT_max')
            
            스_회원_min = mm.get(f'스마트_{mem_key}_min')
            스_회원_max = mm.get(f'스마트_{mem_key}_max')
            스_FIT_min = mm.get('스마트_FIT_min')
            스_FIT_max = mm.get('스마트_FIT_max')
            
            배_회원_min = mm.get(f'배분가_{mem_key}_min')
            배_회원_max = mm.get(f'배분가_{mem_key}_max')
            배_회원_할_min = mm.get(f'배분가_{mem_key}_할인율_min')
            배_회원_할_max = mm.get(f'배분가_{mem_key}_할인율_max')
            배_FIT_min = mm.get('배분가_FIT_min')
            배_FIT_max = mm.get('배분가_FIT_max')
            배_FIT_할_min = mm.get('배분가_FIT_할인율_min')
            배_FIT_할_max = mm.get('배분가_FIT_할인율_max')
            
            # GAP = FIT배분가 - 회원배분가
            gap_min = safe_int(배_FIT_min) - safe_int(배_회원_max) if 배_FIT_min and 배_회원_max else None
            gap_max = safe_int(배_FIT_max) - safe_int(배_회원_min) if 배_FIT_max and 배_회원_min else None
            
            gap_rate_min = gap_min / safe_int(배_FIT_min) if gap_min and 배_FIT_min else None
            gap_rate_max = gap_max / safe_int(배_FIT_max) if gap_max and 배_FIT_max else None
            
            vals = [
                dong, rtype, view,
                fmt_range(正_회원_min, 正_회원_max),
                fmt_range(正_FIT_min, 正_FIT_max),
                fmt_range(스_회원_min, 스_회원_max),
                fmt_range(스_FIT_min, 스_FIT_max),
                fmt_range(배_회원_min, 배_회원_max),
                fmt_pct_range(배_회원_할_min, 배_회원_할_max),
                fmt_range(배_FIT_min, 배_FIT_max),
                fmt_pct_range(배_FIT_할_min, 배_FIT_할_max),
                fmt_range(gap_min, gap_max) if gap_min else '-',
                fmt_pct_range(gap_rate_min, gap_rate_max) if gap_rate_min else '-',
            ]
            
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER if ci <= 3 else ALIGN_RIGHT
                cell.border = BORDER_ALL
                if ci == 12 and gap_min and gap_min > 0:
                    cell.fill = FILL_GAP_POS
                elif ci == 12 and gap_min and gap_min < 0:
                    cell.fill = FILL_GAP_NEG
                if ci in (9, 11, 13):
                    cell.font = FONT_PCT
            
            # Alternate row color
            if (row % 2) == 0:
                for ci in range(1, len(vals)+1):
                    if ws.cell(row=row, column=ci).fill == PatternFill():
                        ws.cell(row=row, column=ci).fill = FILL_LIGHT_GRAY
            row += 1
        
        row += 1
        
        # --- Section: 부대시설 배분가 ---
        ws.merge_cells(f'A{row}:T{row}')
        ws[f'A{row}'] = '▶ 부대시설 배분가'
        ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, size=10)
        ws[f'A{row}'].fill = FILL_SECTION
        row += 1
        
        bude_headers = ['구분', '조식뷔페', '할인율', '오션플레이', '할인율', '레전드①', '할인율', '레전드②', '할인율', '웰컴티①', '할인율', '웰컴티②', '할인율']
        for ci, h in enumerate(bude_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='맑은 고딕', bold=True, size=8, color='FFFFFF')
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1
        
        # 부대시설 is same across all rooms in a period, take first record
        first_key = list(mm_data.keys())[0] if mm_data else None
        if first_key:
            mm = mm_data[first_key]
            bude_vals = [
                '배분가(3인)',
                fmt_range(mm.get('배분가_조식_min'), mm.get('배분가_조식_max')),
                fmt_pct_range(mm.get('배분가_조식_할인율_min'), mm.get('배분가_조식_할인율_max')),
                fmt_range(mm.get('배분가_오션_min'), mm.get('배분가_오션_max')),
                fmt_pct_range(mm.get('배분가_오션_할인율_min'), mm.get('배분가_오션_할인율_max')),
                fmt_range(mm.get('배분가_레전드1_min'), mm.get('배분가_레전드1_max')),
                fmt_pct_range(mm.get('배분가_레전드1_할인율_min'), mm.get('배분가_레전드1_할인율_max')),
                fmt_range(mm.get('배분가_레전드2_min'), mm.get('배분가_레전드2_max')),
                fmt_pct_range(mm.get('배분가_레전드2_할인율_min'), mm.get('배분가_레전드2_할인율_max')),
                fmt_range(mm.get('배분가_웰컴1_min'), mm.get('배분가_웰컴1_max')),
                fmt_pct_range(mm.get('배분가_웰컴1_할인율_min'), mm.get('배분가_웰컴1_할인율_max')),
                fmt_range(mm.get('배분가_웰컴2_min'), mm.get('배분가_웰컴2_max')),
                fmt_pct_range(mm.get('배분가_웰컴2_할인율_min'), mm.get('배분가_웰컴2_할인율_max')),
            ]
            for ci, v in enumerate(bude_vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = FONT_DATA if ci % 2 == 0 else FONT_PCT if '할인율' in str(bude_headers[ci-1]) else FONT_DATA
                cell.alignment = ALIGN_CENTER
                cell.border = BORDER_ALL
            row += 1
        
        row += 1
        
        # --- Section: 판매가 (상품①②) ---
        ws.merge_cells(f'A{row}:T{row}')
        ws[f'A{row}'] = '▶ 판매가 (상품①②)'
        ws[f'A{row}'].font = Font(name='맑은 고딕', bold=True, size=10)
        ws[f'A{row}'].fill = FILL_SECTION
        row += 1
        
        sale_headers = ['동', '타입', '뷰',
                       f'상품① 회원({mem_key})', '상품① FIT', '상품① GAP',
                       '상품① 0.11수수료', '상품① 0.11입금가', '상품① 0.13수수료', '상품① 0.13입금가',
                       f'상품② 회원({mem_key})', '상품② FIT', '상품② GAP',
                       '상품② 0.11수수료', '상품② 0.11입금가', '상품② 0.13수수료', '상품② 0.13입금가']
        for ci, h in enumerate(sale_headers, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = Font(name='맑은 고딕', bold=True, size=7, color='FFFFFF')
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row += 1
        
        for dong, rtype, view in all_rooms:
            key = (dong, rtype, view)
            if key not in mm_data:
                continue
            mm = mm_data[key]
            
            s1_mem_min = mm.get(f'상품1_{mem_key}_min')
            s1_mem_max = mm.get(f'상품1_{mem_key}_max')
            s1_fit_min = mm.get('상품1_FIT_min')
            s1_fit_max = mm.get('상품1_FIT_max')
            s1_gap_min = safe_int(s1_fit_min) - safe_int(s1_mem_max) if s1_fit_min and s1_mem_max else None
            s1_gap_max = safe_int(s1_fit_max) - safe_int(s1_mem_min) if s1_fit_max and s1_mem_min else None
            
            s2_mem_min = mm.get(f'상품2_{mem_key}_min')
            s2_mem_max = mm.get(f'상품2_{mem_key}_max')
            s2_fit_min = mm.get('상품2_FIT_min')
            s2_fit_max = mm.get('상품2_FIT_max')
            s2_gap_min = safe_int(s2_fit_min) - safe_int(s2_mem_max) if s2_fit_min and s2_mem_max else None
            s2_gap_max = safe_int(s2_fit_max) - safe_int(s2_mem_min) if s2_fit_max and s2_mem_min else None
            
            vals = [
                dong, rtype, view,
                fmt_range(s1_mem_min, s1_mem_max), fmt_range(s1_fit_min, s1_fit_max),
                fmt_range(s1_gap_min, s1_gap_max) if s1_gap_min else '-',
                fmt_range(mm.get('상품1_수수료011_min'), mm.get('상품1_수수료011_max')),
                fmt_range(mm.get('상품1_입금가011_min'), mm.get('상품1_입금가011_max')),
                fmt_range(mm.get('상품1_수수료013_min'), mm.get('상품1_수수료013_max')),
                fmt_range(mm.get('상품1_입금가013_min'), mm.get('상품1_입금가013_max')),
                fmt_range(s2_mem_min, s2_mem_max), fmt_range(s2_fit_min, s2_fit_max),
                fmt_range(s2_gap_min, s2_gap_max) if s2_gap_min else '-',
                fmt_range(mm.get('상품2_수수료011_min'), mm.get('상품2_수수료011_max')),
                fmt_range(mm.get('상품2_입금가011_min'), mm.get('상품2_입금가011_max')),
                fmt_range(mm.get('상품2_수수료013_min'), mm.get('상품2_수수료013_max')),
                fmt_range(mm.get('상품2_입금가013_min'), mm.get('상품2_입금가013_max')),
            ]
            
            for ci, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=ci, value=v)
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER if ci <= 3 else ALIGN_RIGHT
                cell.border = BORDER_ALL
                if ci in (6, 13) and s1_gap_min and s1_gap_min > 0:
                    cell.fill = FILL_GAP_POS
            
            if (row % 2) == 0:
                for ci in range(1, len(vals)+1):
                    if ws.cell(row=row, column=ci).fill == PatternFill():
                        ws.cell(row=row, column=ci).fill = FILL_LIGHT_GRAY
            row += 1
        
        row += 2
    
    # ---- Sheet 2: 시즌별 상세 ----
    ws2 = wb.create_sheet('시즌별상세')
    row2 = 1
    ws2.merge_cells(f'A{row2}:R{row2}')
    ws2[f'A{row2}'] = f'■ 소노벨 단양 — 시즌별 상세 요금 ({mode} 기준)'
    ws2[f'A{row2}'].font = FONT_TITLE
    row2 += 2
    
    all_data_sets = [
        ('6월', data_6),
        ('7/1~7/16', data_7),
        ('성수기', data_ss),
    ]
    
    for period_name, records in all_data_sets:
        ws2.merge_cells(f'A{row2}:R{row2}')
        ws2[f'A{row2}'] = period_name
        ws2[f'A{row2}'].font = Font(name='맑은 고딕', bold=True, size=11, color='FFFFFF')
        ws2[f'A{row2}'].fill = PatternFill('solid', fgColor='2F5496')
        row2 += 1
        
        det_headers = ['동', '타입', '뷰', '시즌',
                      f'정상가\n회원({mem_key})', '정상가\nFIT',
                      f'배분가\n회원({mem_key})', f'배분가\n할인율', '배분가\nFIT', '배분가\nFIT할인율',
                      'GAP',
                      f'상품① 회원({mem_key})', '상품① FIT',
                      f'상품② 회원({mem_key})', '상품② FIT']
        
        for ci, h in enumerate(det_headers, 1):
            cell = ws2.cell(row=row2, column=ci, value=h)
            cell.font = Font(name='맑은 고딕', bold=True, size=8, color='FFFFFF')
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_ALL
        row2 += 1
        
        for rec in records:
            正_mem = rec.get(f'정상가_{mem_key}')
            正_fit = rec.get('정상가_FIT')
            배_mem = rec.get(f'배분가_{mem_key}')
            배_mem_h = rec.get(f'배분가_{mem_key}_할인율')
            배_fit = rec.get('배분가_FIT')
            배_fit_h = rec.get('배분가_FIT_할인율')
            gap = safe_int(배_fit) - safe_int(배_mem) if 배_fit and 배_mem else None
            s1_mem = rec.get(f'상품1_{mem_key}')
            s1_fit = rec.get('상품1_FIT')
            s2_mem = rec.get(f'상품2_{mem_key}')
            s2_fit = rec.get('상품2_FIT')
            
            vals = [
                rec['dong'], rec['type'], rec['view'], rec['season'],
                f"{int(正_mem):,}" if 正_mem else '-',
                f"{int(正_fit):,}" if 正_fit else '-',
                f"{int(배_mem):,}" if 배_mem else '-',
                f"{round(배_mem_h*100,1)}%" if 배_mem_h and isinstance(배_mem_h, (int,float)) else '-',
                f"{int(배_fit):,}" if 배_fit else '-',
                f"{round(배_fit_h*100,1)}%" if 배_fit_h and isinstance(배_fit_h, (int,float)) else '-',
                f"{int(gap):,}" if gap else '-',
                f"{int(s1_mem):,}" if s1_mem else '-',
                f"{int(s1_fit):,}" if s1_fit else '-',
                f"{int(s2_mem):,}" if s2_mem else '-',
                f"{int(s2_fit):,}" if s2_fit else '-',
            ]
            
            for ci, v in enumerate(vals, 1):
                cell = ws2.cell(row=row2, column=ci, value=v)
                cell.font = FONT_DATA
                cell.alignment = ALIGN_CENTER if ci <= 4 else ALIGN_RIGHT
                cell.border = BORDER_ALL
                if ci == 11 and gap and gap > 0:
                    cell.fill = FILL_GAP_POS
            
            if (row2 % 2) == 0:
                for ci in range(1, len(vals)+1):
                    if ws2.cell(row=row2, column=ci).fill == PatternFill():
                        ws2.cell(row=row2, column=ci).fill = FILL_LIGHT_GRAY
            row2 += 1
        
        row2 += 1
    
    # ---- Column widths ----
    for ws_obj in [ws, ws2]:
        ws_obj.column_dimensions['A'].width = 7
        ws_obj.column_dimensions['B'].width = 8
        ws_obj.column_dimensions['C'].width = 12
        for c in range(4, 21):
            ws_obj.column_dimensions[get_column_letter(c)].width = 18
    
    # Save
    wb.save(filename)
    print(f"Saved: {filename}")

# Generate both files
create_proposal('무기명', '/sessions/eloquent-laughing-bohr/mnt/sono-competitor-crawler/소노벨단양_6월프로모션_무기명기준.xlsx')
create_proposal('기명', '/sessions/eloquent-laughing-bohr/mnt/sono-competitor-crawler/소노벨단양_6월프로모션_기명기준.xlsx')
