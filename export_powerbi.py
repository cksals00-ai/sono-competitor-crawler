"""
Power BI 연동용 데이터 내보내기
- 날짜별 Excel/CSV 저장
- Power BI가 읽는 최신 파일 덮어쓰기 (sono_competitor_prices_latest.xlsx)
"""

import logging
import os
from datetime import datetime
from pathlib import Path

import pandas as pd
import yaml

logger = logging.getLogger(__name__)


def load_output_config(config_path: str = "config.yaml") -> dict:
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    return cfg.get("output", {})


def export_all(df: pd.DataFrame, config_path: str = "config.yaml"):
    """DataFrame을 Excel + CSV로 저장"""
    out_cfg = load_output_config(config_path)
    export_dir = Path(out_cfg.get("export_dir", "./exports"))
    export_dir.mkdir(parents=True, exist_ok=True)

    today = datetime.today().strftime("%Y%m%d")

    # 날짜별 파일 저장
    excel_name = out_cfg.get("excel_filename", "sono_competitor_prices_{date}.xlsx").format(date=today)
    csv_name = out_cfg.get("csv_filename", "sono_competitor_prices_{date}.csv").format(date=today)

    excel_path = export_dir / excel_name
    csv_path = export_dir / csv_name

    _save_excel(df, excel_path)
    _save_csv(df, csv_path)

    # Power BI용 최신 파일 (항상 덮어쓰기)
    latest_name = out_cfg.get("powerbi_filename", "sono_competitor_prices_latest.xlsx")
    latest_path = export_dir / latest_name
    _save_excel(df, latest_path, sheet_name="최신데이터")
    logger.info(f"Power BI 최신 파일 갱신: {latest_path}")


def _save_excel(df: pd.DataFrame, path: Path, sheet_name: str = "가격데이터"):
    """Excel 저장 (워크시트 포맷 적용)"""
    df_export = _prepare_df(df)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_export.to_excel(writer, index=False, sheet_name=sheet_name)
        _format_excel(writer, sheet_name, df_export)
    logger.info(f"Excel 저장: {path} ({len(df_export)} 행)")


def _save_csv(df: pd.DataFrame, path: Path):
    """CSV 저장 (UTF-8 BOM — Excel 한글 호환)"""
    df_export = _prepare_df(df)
    df_export.to_csv(path, index=False, encoding="utf-8-sig")
    logger.info(f"CSV 저장: {path} ({len(df_export)} 행)")


def _prepare_df(df: pd.DataFrame) -> pd.DataFrame:
    """Power BI 친화적 컬럼명으로 변환"""
    col_map = {
        "crawled_at": "수집일시",
        "property_name": "소노사업장",
        "property_id": "사업장ID",
        "competitor_name": "경쟁사명",
        "ota": "OTA",
        "checkin_date": "체크인",
        "checkout_date": "체크아웃",
        "room_type": "객실유형",
        "price": "판매가(원)",
        "currency": "통화",
        "availability": "판매상태",
        "url": "URL",
        "error": "오류",
        "review_score": "별점(10점)",
        "review_count": "리뷰수",
    }
    df_out = df.copy()
    df_out.rename(columns={k: v for k, v in col_map.items() if k in df_out.columns}, inplace=True)

    # 가격 0 → None 처리 (Power BI에서 공백으로 표시)
    if "판매가(원)" in df_out.columns:
        df_out["판매가(원)"] = df_out["판매가(원)"].replace(0, None)

    return df_out


def _format_excel(writer, sheet_name: str, df: pd.DataFrame):
    """Excel 셀 너비 자동 조정 및 헤더 굵게"""
    try:
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet_name]

        # 헤더 스타일
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 열 너비 자동 조정
        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() or 0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

        # 가격 열 숫자 형식
        price_col = None
        for col_idx, col_name in enumerate(df.columns, start=1):
            if "판매가" in str(col_name):
                price_col = get_column_letter(col_idx)
                break
        if price_col:
            for row in ws.iter_rows(min_row=2, min_col=df.columns.tolist().index("판매가(원)") + 1,
                                    max_col=df.columns.tolist().index("판매가(원)") + 1):
                for cell in row:
                    cell.number_format = "#,##0"

    except Exception as e:
        logger.warning(f"Excel 포맷 적용 실패 (무시): {e}")


if __name__ == "__main__":
    # 테스트: 샘플 데이터로 내보내기 확인
    sample = pd.DataFrame([{
        "crawled_at": "2026-04-15 07:00:00",
        "property_name": "소노벨 비발디파크",
        "property_id": "vivaldi",
        "competitor_name": "엘리시안 강촌",
        "ota": "야놀자",
        "checkin_date": "2026-04-20",
        "checkout_date": "2026-04-21",
        "room_type": "스탠다드 더블",
        "price": 180000,
        "currency": "KRW",
        "availability": "available",
        "url": "",
        "error": "",
    }])
    export_all(sample)
    print("테스트 내보내기 완료. exports/ 폴더를 확인하세요.")
