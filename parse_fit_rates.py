"""
parse_fit_rates.py
──────────────────
소노 자사 FIT 요금 Excel → fit_rates.json 변환기.

Usage:
    python3 parse_fit_rates.py                         # 기본 경로 사용
    python3 parse_fit_rates.py data/fit_rates_source.xlsx

출력: fit_rates.json
"""

from __future__ import annotations

import json
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl

# ── 경로 설정 ─────────────────────────────────────────────────────────────────
DEFAULT_XLSX = Path(__file__).parent / "data" / "fit_rates_source.xlsx"
OUTPUT_JSON  = Path(__file__).parent / "fit_rates.json"

# ── 건너뛸 시트 ───────────────────────────────────────────────────────────────
SKIP_SHEETS = {"공문", "펫프렌들리안내", "취소위약규정(필독)", "GS"}

# ── Excel 시트명 → config.yaml property name(들) 매핑 ────────────────────────
# 한 시트가 여러 카드에 공유될 때는 리스트로 나열 (예: 델피노)
SHEET_TO_PROPS: dict[str, list[str]] = {
    "캄비발디":   ["소노캄 비발디파크"],
    "펫비발디":   ["소노펫 비발디파크"],
    "벨비발디":   ["소노벨 비발디파크"],
    "빌리지비발디": ["빌리지비발디파크"],
    "펠리체비발디": ["소노펠리체 비발디파크"],
    "단양 ":     ["소노벨 단양"],          # 시트명 뒤에 공백 있음
    "단양":      ["소노벨 단양"],
    "청송":      ["소노벨 청송"],
    "천안":      ["소노벨 천안"],
    "변산":      ["소노벨 변산"],
    "고양":      ["소노캄 고양"],
    "벨제주":    ["소노벨 제주"],
    "캄제주":    ["소노캄 제주"],
    "남해":      ["소노벨 남해"],
    "델피노":    ["소노벨 델피노", "소노캄 델피노"],  # 두 카드 공유
    "양양":      ["쏠비치 양양"],
    "삼척":      ["쏠비치 삼척"],
    "거제":      ["소노캄 거제"],
    "진도":      ["쏠비치 진도"],
    "여수":      ["소노캄 여수"],
    "양평":      ["소노벨 양평"],
    "경주":      ["소노캄 경주"],
}


def _parse_sheet(ws) -> dict | None:
    """
    한 시트를 파싱해 아래 dict 반환:
    {
      "room_types": [str, ...],
      "rates": [
        {"date": "YYYY-MM-DD", "요일": str, "시즌명": str,
         "rooms": {room_type: price_int, ...}},
        ...
      ]
    }
    오늘 이후 날짜만 포함. 가격이 모두 0인 행 제외.
    """
    today = date.today()

    # 3행: 객실타입 헤더 (H=col8 부터)
    room_types: list[str] = []
    for c in range(8, ws.max_column + 1):
        v = ws.cell(3, c).value
        if v and str(v).strip():
            room_types.append(str(v).strip())
        else:
            break

    if not room_types:
        return None

    rates: list[dict] = []
    for r in range(5, ws.max_row + 1):
        raw_date = ws.cell(r, 2).value   # B열
        if not raw_date:
            break

        # 날짜 변환
        if isinstance(raw_date, datetime):
            row_date = raw_date.date()
        elif isinstance(raw_date, date):
            row_date = raw_date
        else:
            try:
                row_date = datetime.strptime(str(raw_date)[:10], "%Y-%m-%d").date()
            except Exception:
                continue

        if row_date < today:
            continue

        요일   = str(ws.cell(r, 3).value or "").strip()
        시즌명 = str(ws.cell(r, 6).value or "").strip()

        rooms: dict[str, int] = {}
        for i, rt in enumerate(room_types):
            val = ws.cell(r, 8 + i).value
            try:
                price = int(val or 0)
            except (ValueError, TypeError):
                price = 0
            if price > 0:
                rooms[rt] = price

        if not rooms:
            continue

        rates.append({
            "date":   row_date.isoformat(),
            "요일":   요일,
            "시즌명": 시즌명,
            "rooms":  rooms,
        })

    if not rates:
        return None

    return {"room_types": room_types, "rates": rates}


def parse(xlsx_path: str | Path = DEFAULT_XLSX) -> dict:
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Excel 파일 없음: {xlsx_path}")

    print(f"읽는 중: {xlsx_path}")
    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)

    result: dict[str, dict] = {}

    for sname in wb.sheetnames:
        if sname in SKIP_SHEETS:
            continue

        prop_names = SHEET_TO_PROPS.get(sname)
        if not prop_names:
            print(f"  [건너뜀] 시트 '{sname}' — 매핑 없음")
            continue

        ws = wb[sname]
        parsed = _parse_sheet(ws)
        if not parsed:
            print(f"  [건너뜀] 시트 '{sname}' — 데이터 없음")
            continue

        for pname in prop_names:
            result[pname] = parsed
            print(f"  ✓ {sname:12} → {pname} | 객실타입 {len(parsed['room_types'])}개 | {len(parsed['rates'])}일")

    output = {
        "generated_at": date.today().isoformat(),
        "properties": result,
    }
    return output


def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    data = parse(xlsx_path)

    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"\n저장 완료: {OUTPUT_JSON}")
    print(f"사업장 수: {len(data['properties'])}")


if __name__ == "__main__":
    main()
