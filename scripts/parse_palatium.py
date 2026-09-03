#!/usr/bin/env python3
"""
팔라티움 해운대 바이 소노펠리체 Excel 파서 — BI Python 스크립트(01_PowerBI_데이터가져오기.py)와 동일 로직
- classify_segment / get_channel_name / classify_fit_channel 함수 동일
- rows 배열 출력 → HTML 클라이언트 사이드 필터링
- VAT 1.1 이미 제외 완료 상태 그대로 사용
"""
import json, sys, glob, os, calendar
from datetime import datetime, timedelta
import openpyxl
import pandas as pd

# 팔라티움 해운대 by sonofelice 완전개관 객실수 (6월~ 약 238실/일)
FULL_ROOMS = 238
# 월별 실제 가용 객실박(room-nights) — 다올비전 PMS 월간 실적의 RevPar 역산값(RoomRev/RevPar).
# 단계 개관(ramp-up): 2월 107·3월 149·4월 185·5월 219·6월 238실/일.
# 스냅샷 없는 월(1·7~12월 등)은 FULL_ROOMS(238)×해당 월 일수로 보정.
AVAIL_RN_OVERRIDE = {2: 2996, 3: 4618, 4: 5559, 5: 6802, 6: 7140}
YEAR = 2026
REV_UPLIFT = 1.01  # 사업계획 매출 × 1.01 = 최종 목표 (수수료 1% 가산)
DEFAULT_TARGETS = {  # 사업계획 Excel 부재 시 fallback
    "revenue": 4_000_000_000,
    "rn":      13_200,
    "adr":     300_000,
    "occ":     0.80,
    "revpar":  222_000,
}
BUSINESS_PLAN_PATTERNS = ["*사업계획*.xlsx", "*business_plan*.xlsx", "*BusinessPlan*.xlsx"]
VALID_STATUSES = {"Checked Out","Reservation","In House","Assigned Room","Holding Check Out"}
OVERSEAS_OTA = {"아고다","익스피디아","트립닷컴","부킹닷컴"}
DOMESTIC_OTA = {"놀유니버스","여기어때","타이드스퀘어투어비스","웹투어"}
# ── 요금타입 프로모션 자동 판별 (allowlist 방식) ──────────────────────────
# 표준(비프로모) 요금 화이트리스트. 여기에 없는 유효 요금타입은 "프로모션"으로 자동 판정하여
#   ① 요금타입을 FIT로 흡수(세그먼트/피벗 집계)  ② 원래 요금명을 promo 필드로 보존('프로모션 실적' 별도 집계)
# 새 프로모(익스피디아 캠페인·트립닷컴 아울렛 등)가 등장해도 하드코딩 없이 자동 처리된다.
# 신규 요금타입은 parse 결과 `_new_rates`로 표시 → 스킬(palatium-promo-review)로 표준/프로모 확정.
STANDARD_RATES = {
    "FIT", "소노회원(분양)", "D-멤버스(온라인)", "회원COMP", "Complimentary",
    "Walk-In", "팔라티움(분양회원)", "Direct Call", "House Use", "Rack Rate",
    "팔라티움(임직원)", "소노(임직원)", "기업제휴", "소노기업제휴",
}
# 확정 프로모(검토 완료). 여기에도 표준에도 없는 요금 = 미검토 신규 → _new_rates 알림.
KNOWN_PROMOS = {
    "트립닷컴 동부산 아울렛", "트립비토즈 (8%할인 프로모션)", "익스피다아 visa 캠페인 Room Only 10%",
}
# 프로모션 표시명 정리(선택). 없으면 원래 요금명 그대로 사용.
PROMO_LABEL = {"트립닷컴 동부산 아울렛": "트립닷컴 동부산아울렛"}


def is_promo_rate(rt) -> bool:
    """표준 요금 화이트리스트에 없는 유효 요금타입 = 프로모션."""
    s = str(rt or "").strip()
    if not s or s.lower() == "nan":
        return False
    # 인바운드(전략/일반 여행사)는 프로모션이 아니라 독립 '인바운드' 세그먼트 → FIT 흡수 금지
    if "인바운드" in s:
        return False
    return s not in STANDARD_RATES


def _is_reservation_xlsx(path):
    """첫 시트 앞부분에 '도착일자' 헤더가 있으면 예약정보조회 export로 판단."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for ws in wb.worksheets:
            if not (ws.max_row and ws.max_row > 1):
                continue
            for row in ws.iter_rows(values_only=True, max_row=6):
                if row and any(str(c).strip() == "도착일자" for c in row if c is not None):
                    wb.close(); return True
            break
        wb.close()
    except Exception:
        pass
    return False


def find_excel(data_dir):
    """예약정보조회 export 전체 탐색 (리스트).

    다올비전이 파일명을 날짜/파트(예: 2026010701.xlsx)로 바꿔 내보내도 잡히도록,
    이름 규칙에 의존하지 않고 '도착일자' 컬럼을 가진 xlsx를 모두 반환한다.
    사업계획·임시(~$) 파일은 제외."""
    cands = sorted(set(glob.glob(f"{data_dir}/**/*.xlsx", recursive=True) +
                       glob.glob(f"{data_dir}/*.xlsx")),
                   key=os.path.getmtime, reverse=True)
    # 1차: 이름 힌트(빠름) — 예약정보조회/p_data/palatium
    named = [f for f in cands
             if ("사업계획" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")
                 and any(k in os.path.basename(f).lower() for k in ("예약정보조회", "p_data", "palatium")))]
    # 2차: 이름과 무관하게 내용으로 판별(개명된 파트 포함)
    content = [f for f in cands
               if ("사업계획" not in os.path.basename(f) and not os.path.basename(f).startswith("~$")
                   and _is_reservation_xlsx(f))]
    hits = sorted(set(named) | set(content), key=os.path.getmtime, reverse=True)
    if hits:
        return hits
    raise FileNotFoundError(f"{data_dir}/ 에서 팔라티움 예약정보조회 Excel 없음")


def _find_business_plan(data_dir: str):
    """사업계획 Excel 탐색 (재귀)"""
    for pat in BUSINESS_PLAN_PATTERNS:
        hits = sorted(
            glob.glob(f"{data_dir}/**/{pat}", recursive=True) +
            glob.glob(f"{data_dir}/{pat}"),
            key=os.path.getmtime, reverse=True,
        )
        if hits:
            return hits[0]
    return None


def _load_room_plan(path: str) -> dict:
    """(심사) 객실계획 초안 — 첫 시트(팔라티움 해운대)의 월별/연간 Grand Total(2026 Budget) 파싱.
    컬럼(0-idx): RN=5, ADR=7(천원), Revenue=9(천원). 월 라벨=col2 'N월', 연간='년 간'.
    반환 rev/adr 단위 = 천원 (build 측에서 ×1000 환산)."""
    import re
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    monthly, annual, last = {}, None, None
    for r in ws.iter_rows(values_only=True):
        if not r:
            continue
        c2 = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        c3 = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        if c2:
            last = c2
        if c3 == "Grand Total" and len(r) > 9:
            rec = {"rn": r[5], "adr": r[7], "rev": r[9]}
            mm = re.match(r"^(\d{1,2})\s*월$", last or "")
            if mm:
                monthly[int(mm.group(1))] = rec
            elif last and last.replace(" ", "").startswith("년"):  # '년 간'
                annual = rec
    wb.close()
    if len(monthly) < 12 or not annual:
        raise ValueError(f"{path}: 객실계획 파싱 불완전(months={len(monthly)}, annual={bool(annual)})")
    return {"annual": annual, "monthly": monthly, "source": os.path.basename(path)}


def _load_business_plan(path: str) -> dict:
    """요약 시트 R7~R10 → 월별/연간 사업계획 목표 (천원/실 단위는 시트 그대로)"""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if "요약" not in wb.sheetnames:
        wb.close()
        raise ValueError(f"{path}: '요약' 시트 없음")
    ws = wb["요약"]
    annual = {
        "rn":  ws.cell(7, 4).value,    # D7: 판매객실수(실) 합계
        "occ": ws.cell(8, 4).value,    # D8: 투숙률
        "adr": ws.cell(9, 4).value,    # D9: ADR(천원)
        "rev": ws.cell(10, 4).value,   # D10: 매출액(천원)
    }
    monthly = {}
    for m in range(1, 13):
        col = 4 + m  # E=5 → 1월
        monthly[m] = {
            "rn":  ws.cell(7, col).value,
            "occ": ws.cell(8, col).value,
            "adr": ws.cell(9, col).value,
            "rev": ws.cell(10, col).value,
        }
    wb.close()
    return {"annual": annual, "monthly": monthly, "source": os.path.basename(path)}


def _build_targets(plan: dict | None) -> tuple[dict, dict, str | None]:
    """사업계획 → targets / monthly_targets. 매출에만 REV_UPLIFT(1.01) 적용."""
    if plan is None:
        default_monthly = {
            str(m): {
                "rev": round(DEFAULT_TARGETS["revenue"] / 12),
                "rn":  round(DEFAULT_TARGETS["rn"] / 12),
            } for m in range(1, 13)
        }
        return DEFAULT_TARGETS, default_monthly, None

    a = plan["annual"]
    m = plan["monthly"]
    annual_rev = round(float(a["rev"]) * 1000 * REV_UPLIFT)  # 천원→원, 수수료 1% 가산
    annual_rn  = int(round(float(a["rn"])))
    annual_adr = int(round(float(a["adr"]) * 1000))           # 천원→원 (수수료 미적용)
    annual_occ = round(float(a["occ"]), 4)
    plan_avail = round(annual_rn / annual_occ) if annual_occ > 0 else 0
    annual_revpar = round(annual_rev / plan_avail) if plan_avail else 0

    targets = {
        "revenue": annual_rev,
        "rn":      annual_rn,
        "adr":     annual_adr,
        "occ":     annual_occ,
        "revpar":  annual_revpar,
    }
    monthly_targets = {
        str(mm): {
            "rev": round(float(m[mm]["rev"]) * 1000 * REV_UPLIFT),
            "rn":  int(round(float(m[mm]["rn"]))),
        } for mm in range(1, 13)
    }
    return targets, monthly_targets, plan["source"]


def _load_single(path):
    """단일 xlsx → DataFrame"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for s in wb.worksheets:
        if s.max_row and s.max_row > 1:
            ws = s
            break
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    # 타이틀/날짜 행이 앞에 붙은 경우 실제 헤더 행 탐지
    header_idx = 0
    for i, row in enumerate(rows[:5]):
        if "도착일자" in row:
            header_idx = i
            break
    headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[header_idx])]
    return pd.DataFrame(rows[header_idx + 1:], columns=headers)


def _snapshot_key(path):
    """파일의 스냅샷 시점 키(YYYYMMDD) — 파일명 MMDD 우선, 없으면 mtime.
    여러 스냅샷 합산 시 '최신 스냅샷 우선' dedup 정렬 기준 (mtime은 파일 복사 시
    뒤바뀔 수 있어 파일명 날짜를 우선)."""
    digits = "".join(c for c in os.path.basename(path) if c.isdigit())
    if len(digits) >= 4:
        mm, dd = int(digits[:2]), int(digits[2:4])
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            return f"{YEAR}{mm:02d}{dd:02d}"
    return datetime.fromtimestamp(os.path.getmtime(path)).strftime("%Y%m%d")


def load_df(path):
    """단일 파일 또는 여러 스냅샷 파일 합산 로드.

    마감월(과거 투숙)은 예약정보조회 추출의 롤링 윈도우에서 시간이 지나면 빠지므로,
    옛 추출 파일을 DB에 함께 두면 합산 시 보존된다. 같은 예약번호가 여러 스냅샷에
    있으면 '가장 최신 스냅샷' 행만 유지 → 마감월은 옛 파일에서 살리고, 현재·미래월은
    최신 데이터를 사용."""
    if isinstance(path, list):
        dfs = []
        for p in path:
            df = _load_single(p)
            df["_snap"] = _snapshot_key(p)
            dfs.append(df)
            print(f"  로드: {os.path.basename(p)} ({len(df)}행, snap={_snapshot_key(p)})")
        combined = pd.concat(dfs, ignore_index=True)
        before = len(combined)
        if "예약번호" in combined.columns:
            # 스냅샷 오름차순 정렬 후 keep='last' → 예약번호별 최신 스냅샷 행 유지
            combined = combined.sort_values("_snap", kind="stable")
            _key = combined["예약번호"].astype(str).str.strip()
            _has_key = combined["예약번호"].notna() & (_key != "") & (_key.str.lower() != "none")
            # ⚠ 예약번호 공란(단체/블럭 요청 등 — 개별 예약번호 미부여)은 서로 다른 예약이므로
            #   예약번호로 뭉개면 안 된다(과거 인바운드 단체 119→52 유실 버그). 공란 행은
            #   라인 식별 컬럼 조합으로 '스냅샷 중복'만 제거하고 각 블럭은 보존한다.
            keyed = combined[_has_key].drop_duplicates(subset=["예약번호"], keep="last")
            unkeyed = combined[~_has_key]
            _line = [c for c in ["도착일자", "출발일자", "요금타입", "거래처", "객실타입",
                                 "투숙객명", "상태", "객실수", "박수", "총합계", "시장"]
                     if c in unkeyed.columns]
            if _line:
                unkeyed = unkeyed.drop_duplicates(subset=_line, keep="last")
            combined = pd.concat([keyed, unkeyed], ignore_index=True)
        combined = combined.drop(columns=["_snap"], errors="ignore")
        print(f"  합산: {before}행 → {len(combined)}행 "
              f"(중복 {before - len(combined)}건 제거, 최신 스냅샷 우선)")
        return combined.reset_index(drop=True)
    return _load_single(path)


# ── BI 로직 (01_PowerBI_데이터가져오기.py와 동일) ─────────────────────────
def classify_segment(rt: str, mkt: str = "") -> str:
    rt = str(rt or ""); mkt = str(mkt or "")
    if "소노회원" in rt:
        return "소노회원"
    if "D-멤버스" in rt:
        return "D-멤버스"
    # 인바운드(전략/일반 여행사) = 여행사 인바운드 채널 → 독립 세그먼트
    if "인바운드" in rt:
        return "인바운드"
    if rt == "FIT":
        return "FIT(OTA)"
    if any(k in rt for k in ["팔라티움", "Direct Call", "Walk-In", "Rack Rate"]):
        return "홈페이지(다이렉트)"
    # 요금타입 이름만으론 안 잡히는 OTA/여행사 프로모 요금(예: '트립닷컴 동부산 아울렛')은
    # 시장(해외/국내여행사·FIT)으로 판별해 매출로 정상 분류 (과거 '기타'로 누락되던 버그 교정).
    if any(k in mkt for k in ["여행사", "Traveler", "Foreign", "OTA"]):
        return "FIT(OTA)"
    if "WALK" in mkt:
        return "홈페이지(다이렉트)"
    return "기타"


def classify_fit_channel(seg: str, vendor: str) -> str | None:
    if seg != "FIT(OTA)":
        return None
    if vendor in OVERSEAS_OTA:
        return "FIT-해외OTA"
    if vendor in DOMESTIC_OTA:
        return "FIT-국내OTA"
    return "FIT-기타"


def get_channel_name(seg: str, rt: str, vendor: str) -> str:
    if seg in ("소노회원", "D-멤버스", "인바운드"):
        return seg
    if seg == "FIT(OTA)":
        return vendor if vendor else "기타"
    if seg == "홈페이지(다이렉트)":
        if "Direct Call" in rt:
            return "전화예약"
        if "Walk-In" in rt:
            return "워크인"
        return "팔라티움자체"
    return "기타"


def classify_room(rt: str) -> str:
    for kw, cat in [("Superior","슈페리어"),("Deluxe","디럭스"),("Premier","프리미어"),
                    ("Prestige","프레스티지"),("Presidential","프레지덴셜")]:
        if kw in rt:
            return cat
    return "기타"


def classify_view(rt: str) -> str:
    if "Ocean" in rt:
        return "오션뷰"
    if "VF" in rt:
        return "밸리/포레스트뷰"
    return "기타"


# ─────────────────────────────────────────────────────────────────────────────
def parse(data_dir: str = "data") -> dict:
    paths = find_excel(data_dir)
    df = load_df(paths)
    path = paths[0] if isinstance(paths, list) else paths

    # 타입 변환
    for col in ["도착일자","출발일자","등록일시","취소일자"]:
        df[col] = pd.to_datetime(df[col], errors="coerce")
    df["박수"]   = pd.to_numeric(df["박수"],   errors="coerce").fillna(0).astype(int)
    # 비객실 관리행 제외 → 야간분해 팬텀 RN(매출0/저가)으로 ADR 왜곡 방지.
    #   ① '취소수수료방' 등 취소수수료 관리행(객실수 공란·매출0인데 장기 박수로 팬텀 RN 생성)
    #   ② 원본 객실수가 명시적으로 0인 행(판매객실 아님)
    _rooms_raw = pd.to_numeric(df["객실수"], errors="coerce")
    _admin = df["투숙객명"].astype(str).str.contains("취소수수료", na=False)
    df = df[(~_admin) & (_rooms_raw.fillna(1) > 0)].copy()
    df["객실수"] = pd.to_numeric(df["객실수"], errors="coerce").fillna(1).clip(lower=1).astype(int)
    df["총합계"] = pd.to_numeric(df["총합계"], errors="coerce").fillna(0)
    # 추가상품료(패키지 판별용) — 원본에 없으면 0
    if "추가상품료" in df.columns:
        df["추가상품료"] = pd.to_numeric(df["추가상품료"], errors="coerce").fillna(0)
    else:
        df["추가상품료"] = 0
    # 문자열 차원 정리 (pbix 슬라이서/축이 쓰는 원자 컬럼 포함)
    for col in ["요금타입","거래처","상태","투숙객명","객실타입","시장","경로","국적"]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].fillna("").astype(str).str.strip()

    # HOUSE USE(하우스유즈=호텔 내부사용)는 '판매 객실'이 아니므로 배제
    # (PMS Sold/Occupied Rooms·OCC 산정과 동일. 매출 0이라 매출엔 영향 없음, RN·OCC만 정정).
    df = df[~df["시장"].str.contains("HOUSE", case=False, na=False)].copy()

    # 프로모션 자동 판별: 표준 요금이 아니면 프로모로 태깅(원래명 보존) 후 요금타입을 FIT로 흡수.
    _promo_mask = df["요금타입"].apply(is_promo_rate)
    df["프로모션"] = df["요금타입"].where(_promo_mask).apply(
        lambda rt: (PROMO_LABEL.get(str(rt).strip(), str(rt).strip()) if pd.notna(rt) else None))
    # 미검토 신규 요금(표준·확정프로모 어디에도 없음) → 스킬 검토 대상으로 수집
    _new_rates = sorted({str(rt).strip() for rt, m in zip(df["요금타입"], _promo_mask)
                         if m and str(rt).strip() not in KNOWN_PROMOS})
    df.loc[_promo_mask, "요금타입"] = "FIT"   # FIT 실적에 흡수

    # 파생 컬럼
    df["세그먼트"]     = df.apply(lambda r: classify_segment(r["요금타입"], r["시장"]), axis=1)
    df["FIT채널구분"]  = df.apply(lambda r: classify_fit_channel(r["세그먼트"], r["거래처"]), axis=1)
    df["채널명"]       = df.apply(lambda r: get_channel_name(r["세그먼트"], r["요금타입"], r["거래처"]), axis=1)
    df["세그먼트상세"] = df.apply(lambda r: r["FIT채널구분"] if r["FIT채널구분"] else r["세그먼트"], axis=1)
    df["객실대분류"]   = df["객실타입"].apply(classify_room)
    df["뷰타입"]       = df["객실타입"].apply(classify_view)
    df["패키지여부"]   = (df["추가상품료"] > 0).map({True: "패키지", False: "Room Only"})
    df["RN"]           = df["박수"] * df["객실수"].clip(lower=1)
    df["is_valid"]     = df["상태"].isin(VALID_STATUSES)
    df["is_cancel"]    = df["상태"] == "Cancelled Reservation"
    df["도착월"]       = df["도착일자"].dt.month
    df["도착일"]       = df["도착일자"].dt.day
    df["예약월"]       = df["등록일시"].dt.month
    df["예약일자"]     = df["등록일시"].dt.date.apply(lambda x: x.isoformat() if pd.notna(x) else None)
    df["도착년"]       = df["도착일자"].dt.year
    df["투숙일ISO"]    = df["도착일자"].dt.date.apply(lambda x: x.isoformat() if pd.notna(x) else None)
    df["취소일ISO"]    = df["취소일자"].dt.date.apply(lambda x: x.isoformat() if pd.notna(x) else None)
    df["리드타임"]     = (df["도착일자"] - df["등록일시"].dt.normalize()).dt.days
    df.loc[df["리드타임"] < 0, "리드타임"] = None

    # 재방문 판별 (유효예약 + 비기타 게스트 기준)
    guest_cnts = (df[df["is_valid"] & (df["세그먼트"] != "기타")]
                  .groupby("투숙객명")["예약번호"].count())
    repeat_set = set(guest_cnts[guest_cnts > 1].index)
    df["재방문"] = df["투숙객명"].apply(lambda g: 1 if g in repeat_set else 0)

    # 월별 가용 객실박(room-nights) — 실제 인벤토리 기준 (개관 램프업 반영)
    # 2~4월은 PMS 실측 override, 그 외는 201실 × 해당 월 일수
    avail_by_month = {
        str(m): AVAIL_RN_OVERRIDE.get(m, FULL_ROOMS * calendar.monthrange(YEAR, m)[1])
        for m in range(1, 13)
    }

    # 사업계획 Excel 탐색 → targets / monthly_targets 동적 생성
    plan_path = _find_business_plan(data_dir)
    plan = _load_business_plan(plan_path) if plan_path else None
    targets, monthly_targets, plan_src = _build_targets(plan)
    if plan_src:
        print(f"  ✓ 사업계획 적용: {plan_src} (매출 +{(REV_UPLIFT-1)*100:.0f}%)")
    else:
        print(f"  ⚠ 사업계획 Excel 미발견 — 기본 TARGETS 사용")

    # Slim rows 배열 (클라이언트 필터링용).
    # ★ 매출/객실수는 '투숙 야간(night)' 단위로 분해해 각 야간을 해당 월/일에 귀속
    #   (PMS 시장별 실적과 동일한 박 분배 → 월경계 걸친 예약의 월 귀속 정합).
    #   유효예약은 박수만큼 야간행으로 펼치고(첫 야간 fn=1=예약단위 카운트용),
    #   취소/무효는 예약 단위 1행으로 도착월에 귀속.
    rows_out = []
    for _, r in df.iterrows():
        dims = {
            "seg": r["세그먼트"],
            "fit": r["FIT채널구분"] if pd.notna(r["FIT채널구분"]) else None,
            "ch":  r["채널명"],
            "bm":  int(r["예약월"]) if pd.notna(r["예약월"]) else None,
            "bd":  r["예약일자"] if pd.notna(r["예약일자"]) else None,
            "lead":int(r["리드타임"]) if pd.notna(r["리드타임"]) else None,
            "rt":  r["객실대분류"],
            "vw":  r["뷰타입"],
            "rate": r["요금타입"],
            "vd":  r["거래처"] or "기타",
            "rtf": r["객실타입"],
            "nat": r["국적"] or "미상",
            "rte": r["경로"] or "미상",
            "mkt": r["시장"] or "미상",
            "pkg": r["패키지여부"],
            "promo": r["프로모션"] if pd.notna(r["프로모션"]) else None,
        }
        nights = int(r["박수"]) if pd.notna(r["박수"]) else 0
        rooms  = int(r["객실수"]) if pd.notna(r["객실수"]) else 1
        total  = int(r["총합계"])
        arr    = r["도착일자"]
        if bool(r["is_valid"]) and nights >= 1 and pd.notna(arr):
            per = int(round(total / nights))
            for i in range(nights):
                nd = arr + timedelta(days=i)
                rows_out.append({
                    "m": nd.month, "d": nd.day, "y": nd.year,
                    "ad": nd.date().isoformat(),
                    "r": per, "n": rooms,
                    "v": 1, "k": 0, "cd": None,
                    "fn": 1 if i == 0 else 0,
                    **dims,
                })
        else:
            rows_out.append({
                "m":  int(r["도착월"]) if pd.notna(r["도착월"]) else None,
                "d":  int(r["도착일"]) if pd.notna(r["도착일"]) else None,
                "y":  int(r["도착년"]) if pd.notna(r["도착년"]) else None,
                "ad": r["투숙일ISO"] if pd.notna(r["투숙일ISO"]) else None,
                "r":  total, "n": int(r["RN"]),
                "v":  int(r["is_valid"]), "k": int(r["is_cancel"]),
                "cd": r["취소일ISO"] if pd.notna(r["취소일ISO"]) else None,
                "fn": 1,
                **dims,
            })

    bd_series = df["등록일시"].dropna()
    return {
        "generated_at":    datetime.now().strftime("%Y-%m-%d %H:%M"),
        "source_file":     os.path.basename(path),
        "date_range": {
            "min": df["도착일자"].min().strftime("%Y-%m-%d") if pd.notna(df["도착일자"].min()) else "",
            "max": df["도착일자"].max().strftime("%Y-%m-%d") if pd.notna(df["도착일자"].max()) else "",
        },
        "book_date_range": {
            "min": bd_series.min().strftime("%Y-%m-%d") if len(bd_series) else "",
            "max": bd_series.max().strftime("%Y-%m-%d") if len(bd_series) else "",
        },
        "targets":        targets,
        "monthly_rev_target": round(targets["revenue"] / 12),
        "monthly_targets": monthly_targets,
        "business_plan_source": plan_src,
        "avail_by_month": avail_by_month,
        "new_rates":      _new_rates,   # 미검토 신규 요금타입(스킬 검토 대상)
        "rows":           rows_out,
    }


if __name__ == "__main__":
    data_dir = sys.argv[1] if len(sys.argv) > 1 else "data"
    out_path  = sys.argv[2] if len(sys.argv) > 2 else "data/palatium_data.json"
    result = parse(data_dir)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, separators=(",", ":"))
    rows  = result["rows"]
    valid = [r for r in rows if r["v"]]
    rev   = sum(r["r"] for r in valid)
    rn    = sum(r["n"] for r in valid)
    print(f"✓ {out_path}  ({len(rows)}행)")
    print(f"  매출: {rev:,}원  RN: {rn:,}  ADR: {rev//rn if rn else 0:,}")
    print(f"  세그먼트 분포: {{}}")
    from collections import Counter
    for seg, cnt in Counter(r['seg'] for r in valid).most_common():
        print(f"    {seg}: {cnt}건")
