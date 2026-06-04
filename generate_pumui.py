#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Generate 품의 (approval) Excel files for 소노벨 청송 and 소노캄 거제
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import zipfile
import shutil
import os
import re
from collections import defaultdict

# ============================================================
# COMMON STYLES
# ============================================================
FONT_NAME = '맑은 고딕'
FONT_SIZE = 10
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True)
SUBTITLE_FONT = Font(name=FONT_NAME, size=10, italic=True)
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
DATA_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
HEADER_FILL = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
SEASON_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def set_cell(ws, row, col, value, font=None, alignment=None, fill=None, border=None):
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if font:
        cell.font = font
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    return cell


def to_k(val):
    if val is None or val == 0:
        return 0
    return int(round(val / 1000))


def fmt_range(values):
    vals = sorted(set(values))
    if len(vals) == 0:
        return '-'
    if len(vals) == 1:
        return str(vals[0])
    return f"{vals[0]}~{vals[-1]}"


def fmt_pct_range(values):
    pcts = sorted(set(values))
    if len(pcts) == 0:
        return '-'
    if len(pcts) == 1:
        return f"{pcts[0]}%"
    return f"{pcts[0]}~{pcts[-1]}%"


def calc_discount_pct(normal, discounted):
    if normal == 0 or normal is None or discounted is None:
        return 0
    rate = (1 - discounted / normal) * 100
    return int(round(rate))


def fix_xlsx_rels(filepath):
    """Fix workbook.xml.rels: openpyxl writes absolute paths like /xl/worksheets/sheet1.xml.
    Since rels file is in xl/_rels/, the correct relative path is worksheets/sheet1.xml.
    We replace /xl/ prefix with empty (making it relative to the xl/ directory)."""
    temp_path = filepath + '.tmp'
    with zipfile.ZipFile(filepath, 'r') as zin:
        with zipfile.ZipFile(temp_path, 'w') as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == 'xl/_rels/workbook.xml.rels':
                    content = data.decode('utf-8')
                    # /xl/worksheets/sheet1.xml -> worksheets/sheet1.xml
                    content = content.replace('Target="/xl/', 'Target="')
                    data = content.encode('utf-8')
                zout.writestr(item, data)
    shutil.move(temp_path, filepath)


def write_sheet1_header(ws, title):
    set_cell(ws, 1, 1, title, TITLE_FONT, CENTER_ALIGN)
    ws.merge_cells('A1:L1')
    set_cell(ws, 2, 1, '단위: 천원', SUBTITLE_FONT, LEFT_ALIGN)
    
    # Row 4-5 headers (set values BEFORE merging)
    set_cell(ws, 4, 1, '시즌', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, 4, 2, '구분', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, 4, 3, '객실', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, 4, 6, '부대', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, 4, 9, '판매가', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, 4, 12, '비고', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    for c in range(1, 13):
        cell = ws.cell(row=4, column=c)
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGN
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER
    
    sub = {1: '', 2: '', 3: '회원', 4: 'FIT', 5: 'GAP', 6: '회원', 7: 'FIT', 8: 'GAP',
           9: '회원', 10: 'FIT', 11: 'GAP', 12: ''}
    for c, v in sub.items():
        set_cell(ws, 5, c, v, HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    ws.merge_cells('A4:A5')
    ws.merge_cells('B4:B5')
    ws.merge_cells('C4:E4')
    ws.merge_cells('F4:H4')
    ws.merge_cells('I4:K4')
    ws.merge_cells('L4:L5')


def write_season_block(ws, row, season, d):
    # Write season cell value first, then merge
    set_cell(ws, row, 1, season, HEADER_FONT, CENTER_ALIGN, SEASON_FILL, THIN_BORDER)
    ws.merge_cells(start_row=row, start_column=1, end_row=row+2, end_column=1)
    for rr in range(row, row + 3):
        ws.cell(row=rr, column=1).border = THIN_BORDER
    
    # 정상가
    r = row
    data_r1 = [('정상가', CENTER_ALIGN), (d['room_n_m'], RIGHT_ALIGN), (d['room_n_f'], RIGHT_ALIGN), 
                ('-', CENTER_ALIGN), (d['bude_n_m'], RIGHT_ALIGN), (d['bude_n_f'], RIGHT_ALIGN),
                ('-', CENTER_ALIGN), (d['sale_n_m'], RIGHT_ALIGN), (d['sale_n_f'], RIGHT_ALIGN),
                ('-', CENTER_ALIGN), ('-', CENTER_ALIGN)]
    for i, (v, al) in enumerate(data_r1):
        set_cell(ws, r, 2+i, v, DATA_FONT, al, None, THIN_BORDER)
    
    # 배분가
    r = row + 1
    data_r2 = [('배분가', CENTER_ALIGN), (d['room_a_m'], RIGHT_ALIGN), (d['room_a_f'], RIGHT_ALIGN),
                (d['room_gap'], RIGHT_ALIGN), (d['bude_a_m'], RIGHT_ALIGN), (d['bude_a_f'], RIGHT_ALIGN),
                (d['bude_gap'], RIGHT_ALIGN), (d['sale_a_m'], RIGHT_ALIGN), (d['sale_a_f'], RIGHT_ALIGN),
                (d['sale_gap'], RIGHT_ALIGN), ('-', CENTER_ALIGN)]
    for i, (v, al) in enumerate(data_r2):
        set_cell(ws, r, 2+i, v, DATA_FONT, al, None, THIN_BORDER)
    
    # 할인율
    r = row + 2
    data_r3 = [('할인율', CENTER_ALIGN), (d['room_d_m'], CENTER_ALIGN), (d['room_d_f'], CENTER_ALIGN),
                ('-', CENTER_ALIGN), (d['bude_d_m'], CENTER_ALIGN), (d['bude_d_f'], CENTER_ALIGN),
                ('-', CENTER_ALIGN), (d['sale_d_m'], CENTER_ALIGN), (d['sale_d_f'], CENTER_ALIGN),
                ('-', CENTER_ALIGN), ('-', CENTER_ALIGN)]
    for i, (v, al) in enumerate(data_r3):
        set_cell(ws, r, 2+i, v, DATA_FONT, al, None, THIN_BORDER)
    
    return row + 3


def write_footer_rows(ws, row):
    set_cell(ws, row, 1, '수수료', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, row, 2, '11~13%', DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
    for c in range(2, 13):
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
    
    row += 1
    set_cell(ws, row, 1, 'KPI', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws, row, 2, '판매목표 달성', DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
    for c in range(2, 13):
        ws.cell(row=row, column=c).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)


# ============================================================
# FILE 1: 소노벨 청송
# ============================================================
def generate_cheongsong():
    UPLOADS = '/sessions/hopeful-relaxed-mccarthy/mnt/uploads'
    OUTPUT = '/sessions/hopeful-relaxed-mccarthy/mnt/sono-competitor-crawler'
    
    src = openpyxl.load_workbook(
        os.path.join(UPLOADS, '5daa5c2f-GS__________________________________2__________.xlsx'),
        data_only=True
    )
    ws_src = src['1박']
    
    def normalize_season(s):
        if s is None:
            return None
        s = str(s).strip()
        if s in ('주 중', '주중'):
            return '주중'
        if s == '금요일':
            return '금'
        if s == '토요일':
            return '토'
        if '하이' in s:
            return '하이'
        if '골드' in s:
            return '골드'
        if '스페셜' in s:
            return '스페셜'
        return s

    all_rows = []
    for r in list(range(22, 31)) + list(range(36, 48)) + list(range(53, 74)):
        season_raw = ws_src.cell(row=r, column=8).value
        if season_raw is None:
            continue
        season = normalize_season(season_raw)
        
        rnm = ws_src.cell(row=r, column=10).value   # J = 무기명정상
        rnf = ws_src.cell(row=r, column=11).value    # K = FIT정상
        ram = ws_src.cell(row=r, column=21).value    # U = 무기명배분
        raf = ws_src.cell(row=r, column=23).value    # W = FIT배분
        
        if rnm is None:
            continue
        
        all_rows.append({
            'season': season,
            'room_normal_member': rnm, 'room_normal_fit': rnf,
            'room_alloc_member': ram, 'room_alloc_fit': raf,
        })
    
    season_order = ['주중', '금', '토', '골드', '하이', '스페셜']
    season_data = defaultdict(list)
    for row in all_rows:
        season_data[row['season']].append(row)
    
    BUDE_NORMAL = 78000 + 40000
    BUDE_ALLOC = 48000 + 20000
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = '배분가'
    write_sheet1_header(ws1, '소노벨 청송 - 배분가')
    
    current_row = 6
    for season in season_order:
        if season not in season_data:
            continue
        rows = season_data[season]
        
        rnm_v = [to_k(r['room_normal_member']) for r in rows]
        rnf_v = [to_k(r['room_normal_fit']) for r in rows]
        ram_v = [to_k(r['room_alloc_member']) for r in rows]
        raf_v = [to_k(r['room_alloc_fit']) for r in rows]
        
        bude_n_k = to_k(BUDE_NORMAL)
        bude_a_k = to_k(BUDE_ALLOC)
        
        snm_v = [v + bude_n_k for v in rnm_v]
        snf_v = [v + bude_n_k for v in rnf_v]
        sam_v = [v + bude_a_k for v in ram_v]
        saf_v = [v + bude_a_k for v in raf_v]
        
        rgap_v = [f - m for f, m in zip(raf_v, ram_v)]
        sgap_v = [f - m for f, m in zip(saf_v, sam_v)]
        
        rdm_v = [calc_discount_pct(r['room_normal_member'], r['room_alloc_member']) for r in rows]
        rdf_v = [calc_discount_pct(r['room_normal_fit'], r['room_alloc_fit']) for r in rows]
        bd = calc_discount_pct(BUDE_NORMAL, BUDE_ALLOC)
        sdm_v = [calc_discount_pct(r['room_normal_member'] + BUDE_NORMAL, r['room_alloc_member'] + BUDE_ALLOC) for r in rows]
        sdf_v = [calc_discount_pct(r['room_normal_fit'] + BUDE_NORMAL, r['room_alloc_fit'] + BUDE_ALLOC) for r in rows]
        
        d = {
            'room_n_m': fmt_range(rnm_v), 'room_n_f': fmt_range(rnf_v),
            'room_a_m': fmt_range(ram_v), 'room_a_f': fmt_range(raf_v),
            'room_gap': fmt_range(rgap_v),
            'bude_n_m': str(bude_n_k), 'bude_n_f': str(bude_n_k),
            'bude_a_m': str(bude_a_k), 'bude_a_f': str(bude_a_k),
            'bude_gap': '0',
            'sale_n_m': fmt_range(snm_v), 'sale_n_f': fmt_range(snf_v),
            'sale_a_m': fmt_range(sam_v), 'sale_a_f': fmt_range(saf_v),
            'sale_gap': fmt_range(sgap_v),
            'room_d_m': fmt_pct_range(rdm_v), 'room_d_f': fmt_pct_range(rdf_v),
            'bude_d_m': f"{bd}%", 'bude_d_f': f"{bd}%",
            'sale_d_m': fmt_pct_range(sdm_v), 'sale_d_f': fmt_pct_range(sdf_v),
        }
        current_row = write_season_block(ws1, current_row, season, d)
    
    write_footer_rows(ws1, current_row)
    
    for c, w in {1: 10, 2: 10, 3: 14, 4: 14, 5: 12, 6: 10, 7: 10, 8: 10, 9: 14, 10: 14, 11: 12, 12: 10}.items():
        ws1.column_dimensions[get_column_letter(c)].width = w
    
    # Sheet 2: 부대시설 상세내용
    ws2 = wb.create_sheet('부대시설 상세내용')
    set_cell(ws2, 1, 1, '소노벨 청송 - 부대시설 상세내용', TITLE_FONT, CENTER_ALIGN)
    ws2.merge_cells('A1:H1')
    set_cell(ws2, 2, 1, '단위: 천원', SUBTITLE_FONT, LEFT_ALIGN)
    
    set_cell(ws2, 4, 1, '시즌', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 2, '구분', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 3, '조식', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 6, '온천(솔샘온천)', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    for c in [4, 5, 7, 8]:
        cell = ws2.cell(row=4, column=c)
        cell.font = HEADER_FONT; cell.alignment = CENTER_ALIGN; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
    
    for c, v in {1: '', 2: '', 3: '회원', 4: 'FIT', 5: 'GAP', 6: '회원', 7: 'FIT', 8: 'GAP'}.items():
        set_cell(ws2, 5, c, v, HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    ws2.merge_cells('A4:A5')
    ws2.merge_cells('B4:B5')
    ws2.merge_cells('C4:E4')
    ws2.merge_cells('F4:H4')
    
    josik_n = to_k(78000); josik_a = to_k(48000); josik_d = calc_discount_pct(78000, 48000)
    oncheon_n = to_k(40000); oncheon_a = to_k(20000); oncheon_d = calc_discount_pct(40000, 20000)
    
    cr = 6
    for bs in ['주중', '주말']:
        set_cell(ws2, cr, 1, bs, HEADER_FONT, CENTER_ALIGN, SEASON_FILL, THIN_BORDER)
        ws2.merge_cells(start_row=cr, start_column=1, end_row=cr+2, end_column=1)
        for rr in range(cr, cr+3):
            ws2.cell(row=rr, column=1).border = THIN_BORDER
        
        for c, v in {2: '정상가', 3: str(josik_n), 4: str(josik_n), 5: '-', 6: str(oncheon_n), 7: str(oncheon_n), 8: '-'}.items():
            al = CENTER_ALIGN if v in ('-', '정상가') else RIGHT_ALIGN
            set_cell(ws2, cr, c, v, DATA_FONT, al, None, THIN_BORDER)
        for c, v in {2: '배분가', 3: str(josik_a), 4: str(josik_a), 5: '0', 6: str(oncheon_a), 7: str(oncheon_a), 8: '0'}.items():
            al = CENTER_ALIGN if v == '배분가' else RIGHT_ALIGN
            set_cell(ws2, cr+1, c, v, DATA_FONT, al, None, THIN_BORDER)
        for c, v in {2: '할인율', 3: f'{josik_d}%', 4: f'{josik_d}%', 5: '-', 6: f'{oncheon_d}%', 7: f'{oncheon_d}%', 8: '-'}.items():
            set_cell(ws2, cr+2, c, v, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
        cr += 3
    
    for c, w in {1: 10, 2: 10, 3: 12, 4: 12, 5: 10, 6: 14, 7: 14, 8: 10}.items():
        ws2.column_dimensions[get_column_letter(c)].width = w
    
    # Sheet 3: 판매&홍보 일정
    ws3 = wb.create_sheet('판매&홍보 일정')
    set_cell(ws3, 1, 1, '소노벨 청송 - 판매&홍보 일정', TITLE_FONT, CENTER_ALIGN)
    ws3.merge_cells('A1:D1')
    
    for c, v in {1: '구분', 2: '기간', 3: '내용', 4: '비고'}.items():
        set_cell(ws3, 3, c, v, HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    sched = [
        ('판매기간', '6.11~6.30', '①6.11~6.14 인플루언서(dew_with_you)\n②6.15~6.21\n③6.22~6.30', '-'),
        ('투숙기간', '6.11~8.22', '객실+조식+솔샘온천+할인권 / 3인', '-'),
    ]
    for i, (a, b, cv, dd) in enumerate(sched):
        r = 4 + i
        set_cell(ws3, r, 1, a, DATA_FONT, CENTER_ALIGN, SEASON_FILL, THIN_BORDER)
        set_cell(ws3, r, 2, b, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
        set_cell(ws3, r, 3, cv, DATA_FONT, LEFT_ALIGN, None, THIN_BORDER)
        set_cell(ws3, r, 4, dd, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
    
    for c, w in {1: 12, 2: 15, 3: 45, 4: 10}.items():
        ws3.column_dimensions[get_column_letter(c)].width = w
    ws3.row_dimensions[4].height = 50
    
    outpath = os.path.join(OUTPUT, '품의_소노벨청송.xlsx')
    wb.save(outpath)
    fix_xlsx_rels(outpath)
    print(f"Created: {outpath}")


# ============================================================
# FILE 2: 소노캄 거제
# ============================================================
def generate_geoje():
    UPLOADS = '/sessions/hopeful-relaxed-mccarthy/mnt/uploads'
    OUTPUT = '/sessions/hopeful-relaxed-mccarthy/mnt/sono-competitor-crawler'
    
    src = openpyxl.load_workbook(
        os.path.join(UPLOADS, '10306ffb-GS_________________________________________6.16________________.xlsx'),
        data_only=True
    )
    ws_src = src['상품①']
    
    def norm_season(s):
        if s is None:
            return None
        s = str(s).strip()
        mapping = {
            '주중': '주중', '금요일': '금', '토요일': '토',
            '골드': '골드', '하이': '하이',
            '골드(금)': '골드금', '골드(토)': '골드토', '스페셜': '스페셜'
        }
        return mapping.get(s, s)
    
    OCEAN_NORMAL = 212000
    LEGEND_NORMAL = 33000
    BUDE_NORMAL = OCEAN_NORMAL + LEGEND_NORMAL
    
    all_rows = []
    for row_range, ocean_alloc, legend_alloc in [
        (range(19, 51), 60000, 12000),
        (range(51, 83), 60000, 12000),
        (range(83, 115), 75000, 12000),
        (range(115, 150), 60000, 12000),
    ]:
        ba = ocean_alloc + legend_alloc
        for r in row_range:
            sr = ws_src.cell(row=r, column=7).value
            if sr is None:
                continue
            season = norm_season(sr)
            rnm = ws_src.cell(row=r, column=9).value
            rnf = ws_src.cell(row=r, column=10).value
            ram = ws_src.cell(row=r, column=21).value
            raf = ws_src.cell(row=r, column=23).value
            if rnm is None:
                continue
            all_rows.append({
                'season': season,
                'room_normal_member': rnm, 'room_normal_fit': rnf,
                'room_alloc_member': ram, 'room_alloc_fit': raf,
                'bude_alloc': ba, 'ocean_alloc': ocean_alloc, 'legend_alloc': legend_alloc,
            })
    
    season_order = ['주중', '금', '토', '골드', '하이', '골드금', '골드토', '스페셜']
    season_data = defaultdict(list)
    for row in all_rows:
        season_data[row['season']].append(row)
    
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = '배분가'
    write_sheet1_header(ws1, '소노캄 거제 - 배분가')
    
    current_row = 6
    for season in season_order:
        if season not in season_data:
            continue
        rows = season_data[season]
        
        rnm_v = [to_k(r['room_normal_member']) for r in rows]
        rnf_v = [to_k(r['room_normal_fit']) for r in rows]
        ram_v = [to_k(r['room_alloc_member']) for r in rows]
        raf_v = [to_k(r['room_alloc_fit']) for r in rows]
        
        bude_n_k = to_k(BUDE_NORMAL)
        bude_a_vals = [to_k(r['bude_alloc']) for r in rows]
        
        snm_v = [to_k(r['room_normal_member']) + bude_n_k for r in rows]
        snf_v = [to_k(r['room_normal_fit']) + bude_n_k for r in rows]
        sam_v = [to_k(r['room_alloc_member']) + to_k(r['bude_alloc']) for r in rows]
        saf_v = [to_k(r['room_alloc_fit']) + to_k(r['bude_alloc']) for r in rows]
        
        rgap_v = [f - m for f, m in zip(raf_v, ram_v)]
        sgap_v = [f - m for f, m in zip(saf_v, sam_v)]
        
        rdm_v = [calc_discount_pct(r['room_normal_member'], r['room_alloc_member']) for r in rows]
        rdf_v = [calc_discount_pct(r['room_normal_fit'], r['room_alloc_fit']) for r in rows]
        bdm_v = [calc_discount_pct(BUDE_NORMAL, r['bude_alloc']) for r in rows]
        sdm_v = [calc_discount_pct(r['room_normal_member'] + BUDE_NORMAL, r['room_alloc_member'] + r['bude_alloc']) for r in rows]
        sdf_v = [calc_discount_pct(r['room_normal_fit'] + BUDE_NORMAL, r['room_alloc_fit'] + r['bude_alloc']) for r in rows]
        
        d = {
            'room_n_m': fmt_range(rnm_v), 'room_n_f': fmt_range(rnf_v),
            'room_a_m': fmt_range(ram_v), 'room_a_f': fmt_range(raf_v),
            'room_gap': fmt_range(rgap_v),
            'bude_n_m': str(bude_n_k), 'bude_n_f': str(bude_n_k),
            'bude_a_m': fmt_range(bude_a_vals), 'bude_a_f': fmt_range(bude_a_vals),
            'bude_gap': '0',
            'sale_n_m': fmt_range(snm_v), 'sale_n_f': fmt_range(snf_v),
            'sale_a_m': fmt_range(sam_v), 'sale_a_f': fmt_range(saf_v),
            'sale_gap': fmt_range(sgap_v),
            'room_d_m': fmt_pct_range(list(set(rdm_v))), 'room_d_f': fmt_pct_range(list(set(rdf_v))),
            'bude_d_m': fmt_pct_range(list(set(bdm_v))), 'bude_d_f': fmt_pct_range(list(set(bdm_v))),
            'sale_d_m': fmt_pct_range(list(set(sdm_v))), 'sale_d_f': fmt_pct_range(list(set(sdf_v))),
        }
        current_row = write_season_block(ws1, current_row, season, d)
    
    write_footer_rows(ws1, current_row)
    
    for c, w in {1: 10, 2: 10, 3: 14, 4: 14, 5: 12, 6: 12, 7: 12, 8: 10, 9: 14, 10: 14, 11: 12, 12: 10}.items():
        ws1.column_dimensions[get_column_letter(c)].width = w
    
    # Sheet 2: 부대시설 상세내용
    ws2 = wb.create_sheet('부대시설 상세내용')
    set_cell(ws2, 1, 1, '소노캄 거제 - 부대시설 상세내용', TITLE_FONT, CENTER_ALIGN)
    ws2.merge_cells('A1:H1')
    set_cell(ws2, 2, 1, '단위: 천원 (3인 기준)', SUBTITLE_FONT, LEFT_ALIGN)
    
    set_cell(ws2, 4, 1, '시즌', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 2, '구분', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 3, '오션(요트)', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    set_cell(ws2, 4, 6, '레전드히어로즈', HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    for c in [4, 5, 7, 8]:
        cell = ws2.cell(row=4, column=c)
        cell.font = HEADER_FONT; cell.alignment = CENTER_ALIGN; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
    
    for c, v in {1: '', 2: '', 3: '회원', 4: 'FIT', 5: 'GAP', 6: '회원', 7: 'FIT', 8: 'GAP'}.items():
        set_cell(ws2, 5, c, v, HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    ws2.merge_cells('A4:A5')
    ws2.merge_cells('B4:B5')
    ws2.merge_cells('C4:E4')
    ws2.merge_cells('F4:H4')
    
    ocean_nk = to_k(OCEAN_NORMAL)
    legend_nk = to_k(LEGEND_NORMAL)
    
    cr = 6
    for bs, oa, la in [('주중', 60, 12), ('주말', 60, 12), ('주중(성수기)', 75, 12), ('주말(성수기)', 75, 12)]:
        set_cell(ws2, cr, 1, bs, HEADER_FONT, CENTER_ALIGN, SEASON_FILL, THIN_BORDER)
        ws2.merge_cells(start_row=cr, start_column=1, end_row=cr+2, end_column=1)
        for rr in range(cr, cr+3):
            ws2.cell(row=rr, column=1).border = THIN_BORDER
        
        od = calc_discount_pct(OCEAN_NORMAL, oa * 1000)
        ld = calc_discount_pct(LEGEND_NORMAL, la * 1000)
        
        for c, v in {2: '정상가', 3: str(ocean_nk), 4: str(ocean_nk), 5: '-', 6: str(legend_nk), 7: str(legend_nk), 8: '-'}.items():
            al = CENTER_ALIGN if v in ('-', '정상가') else RIGHT_ALIGN
            set_cell(ws2, cr, c, v, DATA_FONT, al, None, THIN_BORDER)
        for c, v in {2: '배분가', 3: str(oa), 4: str(oa), 5: '0', 6: str(la), 7: str(la), 8: '0'}.items():
            al = CENTER_ALIGN if v == '배분가' else RIGHT_ALIGN
            set_cell(ws2, cr+1, c, v, DATA_FONT, al, None, THIN_BORDER)
        for c, v in {2: '할인율', 3: f'{od}%', 4: f'{od}%', 5: '-', 6: f'{ld}%', 7: f'{ld}%', 8: '-'}.items():
            set_cell(ws2, cr+2, c, v, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
        cr += 3
    
    for c, w in {1: 14, 2: 10, 3: 12, 4: 12, 5: 10, 6: 14, 7: 14, 8: 10}.items():
        ws2.column_dimensions[get_column_letter(c)].width = w
    
    # Sheet 3: 판매&홍보 일정
    ws3 = wb.create_sheet('판매&홍보 일정')
    set_cell(ws3, 1, 1, '소노캄 거제 - 판매&홍보 일정', TITLE_FONT, CENTER_ALIGN)
    ws3.merge_cells('A1:D1')
    
    for c, v in {1: '구분', 2: '기간', 3: '내용', 4: '비고'}.items():
        set_cell(ws3, 3, c, v, HEADER_FONT, CENTER_ALIGN, HEADER_FILL, THIN_BORDER)
    
    for i, (a, b, cv, dd) in enumerate([
        ('판매기간', '6.16~8.22', '인플루언서(@지운맘) 마케팅', '-'),
        ('투숙기간', '6.16~9.30', '객실+오션 or 요트+레전드히어로즈+할인권', '-'),
    ]):
        r = 4 + i
        set_cell(ws3, r, 1, a, DATA_FONT, CENTER_ALIGN, SEASON_FILL, THIN_BORDER)
        set_cell(ws3, r, 2, b, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
        set_cell(ws3, r, 3, cv, DATA_FONT, LEFT_ALIGN, None, THIN_BORDER)
        set_cell(ws3, r, 4, dd, DATA_FONT, CENTER_ALIGN, None, THIN_BORDER)
    
    for c, w in {1: 12, 2: 15, 3: 45, 4: 10}.items():
        ws3.column_dimensions[get_column_letter(c)].width = w
    
    outpath = os.path.join(OUTPUT, '품의_소노캄거제.xlsx')
    wb.save(outpath)
    fix_xlsx_rels(outpath)
    print(f"Created: {outpath}")


if __name__ == '__main__':
    generate_cheongsong()
    generate_geoje()
    print("Done!")
